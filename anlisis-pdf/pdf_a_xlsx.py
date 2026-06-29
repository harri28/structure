"""
Convierte un presupuesto S10 exportado en PDF a XLSX importable por el sistema.

Uso:
  python pdf_a_xlsx.py <archivo.pdf> [salida.xlsx]

Ejemplo:
  python pdf_a_xlsx.py "../1.00 PRESUPUESTO DE PAVIMENTACION.pdf"
"""
import sys
import os
import re
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Patrón de código de ítem S10: 01  /  01.01  /  01.01.01  /  01.01.01.01
RE_CODIGO = re.compile(r'^\d{1,2}(\.\d{2,})*$')
RE_NUMERO = re.compile(r'^[\d,]+\.\d+$')


def es_codigo(s):
    s = s.strip()
    return bool(RE_CODIGO.match(s))


def es_numero(s):
    s = s.strip().replace(',', '')
    try:
        float(s)
        return True
    except ValueError:
        return False


def extraer_palabras_por_fila(page):
    """
    Extrae palabras con posición y las agrupa por línea (Y similar).
    Devuelve lista de filas, cada fila = lista de (x0, texto).
    """
    words = page.get_text("words")  # (x0,y0,x1,y1,text,block,line,word)
    if not words:
        return []

    # Agrupar por Y con tolerancia de 3 puntos
    filas = {}
    for w in words:
        x0, y0, x1, y1, texto = w[0], w[1], w[2], w[3], w[4]
        y_key = round(y0 / 3) * 3
        filas.setdefault(y_key, []).append((x0, texto))

    resultado = []
    for y in sorted(filas):
        fila = sorted(filas[y], key=lambda t: t[0])  # ordenar por X
        resultado.append(fila)
    return resultado


def parsear_pagina(page):
    """
    Reconstruye partidas de una página S10.
    Columnas S10 aproximadas (puntos PDF):
      Item       ≈ x < 80
      Descripción≈ 80 < x < 380
      Und.       ≈ 380 < x < 430
      Metrado    ≈ 430 < x < 510
      Precio     ≈ 510 < x < 590
      Parcial    ≈ x > 590
    """
    COLS = {
        'item':    (0,    108),
        'desc':    (108,  342),
        'und':     (342,  408),
        'metrado': (408,  452),
        'precio':  (452,  500),
        'parcial': (500,  9999),
    }

    def col_de(x):
        for nombre, (x0, x1) in COLS.items():
            if x0 <= x < x1:
                return nombre
        return 'parcial'

    filas_raw = extraer_palabras_por_fila(page)
    registros = []  # cada registro = dict

    # Saltar encabezado: buscar la fila que contenga "Item" o "Descripción"
    inicio = 0
    for i, fila in enumerate(filas_raw):
        textos = [t for _, t in fila]
        if any(t.lower() in ('item', 'ítem', 'descripción', 'descripcion') for t in textos):
            inicio = i + 1
            break

    for fila in filas_raw[inicio:]:
        row = {'item': [], 'desc': [], 'und': [], 'metrado': [], 'precio': [], 'parcial': []}
        for x, texto in fila:
            col = col_de(x)
            # Números grandes en zona 'und' son en realidad metrado (S10 los desplaza a la izquierda)
            if col == 'und' and es_numero(texto):
                col = 'metrado'
            row[col].append(texto)

        item    = ' '.join(row['item']).strip()
        desc    = ' '.join(row['desc']).strip()
        und     = ' '.join(row['und']).strip()
        metrado = ' '.join(row['metrado']).strip()
        precio  = ' '.join(row['precio']).strip()
        parcial = ' '.join(row['parcial']).strip()

        # Ignorar filas vacías y líneas de totales finales
        if not item and not desc:
            continue
        if desc.upper() in ('COSTO DIRECTO', 'GASTOS GENERALES', 'UTILIDAD',
                             'SUB TOTAL', 'IMPUESTO', 'TOTAL PRESUPUESTO', 'SON'):
            continue
        if 'TOTAL PRESUPUESTO' in desc.upper():
            continue

        registros.append({
            'item':    item,
            'desc':    desc,
            'und':     und,
            'metrado': metrado,
            'precio':  precio,
            'parcial': parcial,
        })

    return registros


def limpiar_numero(s):
    if not s:
        return ''
    s = s.strip().replace(',', '')
    try:
        return float(s)
    except ValueError:
        return ''


def fusionar_continuaciones(registros):
    """
    Fusiona filas de continuación:
    - Si un registro tiene código pero sin descripción, la siguiente fila
      con descripción (y sin código) es continuación.
    - Si un registro no tiene código pero tiene descripción, es continuación
      de la fila anterior.
    """
    resultado = []
    for r in registros:
        item = r['item']
        desc = r['desc']

        if not item and not desc:
            continue

        if not item and desc and resultado:
            # Línea de continuación: fusionar con la fila anterior
            prev = resultado[-1]
            if not prev['desc']:
                # El anterior tenía código pero no descripción
                prev['desc'] = desc
                if not prev['und']:   prev['und']     = r['und']
                if not prev['metrado'] and r['metrado'] != '': prev['metrado'] = limpiar_numero(r['metrado'])
                if not prev['precio']  and r['precio']  != '': prev['precio']  = limpiar_numero(r['precio'])
                if not prev['parcial'] and r['parcial'] != '': prev['parcial'] = limpiar_numero(r['parcial'])
            else:
                # Descripción que desbordó a la siguiente línea — agregar al nombre
                prev['desc'] = prev['desc'] + ' ' + desc
            continue

        resultado.append({
            'item':    item if es_codigo(item) else '',
            'desc':    desc,
            'und':     r['und'],
            'metrado': limpiar_numero(r['metrado']),
            'precio':  limpiar_numero(r['precio']),
            'parcial': limpiar_numero(r['parcial']),
        })

    return resultado


def parsear_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    todos = []

    for num in range(len(doc)):
        page = doc[num]
        registros = parsear_pagina(page)
        todos.extend(registros)

    doc.close()
    return fusionar_continuaciones(todos)


def guardar_xlsx(partidas, salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Presupuesto'

    # Encabezados
    headers = ['Ítem', 'Descripción', 'Und', 'Metrado', 'P.U.', 'Parcial S/.']
    header_font  = Font(bold=True, color='FFFFFF', size=10)
    header_fill  = PatternFill('solid', fgColor='1D4ED8')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 18
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    fill_sec  = PatternFill('solid', fgColor='DBEAFE')
    fill_leaf = PatternFill('solid', fgColor='FFFFFF')
    font_sec  = Font(bold=True, size=9)
    font_leaf = Font(size=9)

    for row_num, p in enumerate(partidas, 2):
        item = p['item']
        es_hoja = item and len(item.split('.')) >= 3 and p['metrado'] != ''

        ws.cell(row=row_num, column=1, value=item)
        ws.cell(row=row_num, column=2, value=p['desc'])
        ws.cell(row=row_num, column=3, value=p['und'])
        ws.cell(row=row_num, column=4, value=p['metrado'] if p['metrado'] != '' else None)
        ws.cell(row=row_num, column=5, value=p['precio']  if p['precio']  != '' else None)
        ws.cell(row=row_num, column=6, value=p['parcial'] if p['parcial'] != '' else None)

        fill = fill_leaf if es_hoja else fill_sec
        font = font_leaf if es_hoja else font_sec
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).fill = fill
            ws.cell(row=row_num, column=col).font = font

    wb.save(salida)
    print(f'XLSX guardado: {salida}')
    print(f'Total filas:   {len(partidas)}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    pdf_path = sys.argv[1]
    if not os.path.exists(pdf_path):
        print(f'Archivo no encontrado: {pdf_path}')
        sys.exit(1)

    base     = os.path.splitext(os.path.basename(pdf_path))[0]
    salida   = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), f'{base}.xlsx'
    )

    print(f'Procesando: {pdf_path}')
    partidas = parsear_pdf(pdf_path)

    print('\n=== VISTA PREVIA (primeras 20 filas) ===')
    for p in partidas[:20]:
        print(f"  {p['item']:<18} {p['desc'][:50]:<50} {p['und']:<8} {str(p['metrado']):<12} {str(p['precio']):<12}")

    guardar_xlsx(partidas, salida)
