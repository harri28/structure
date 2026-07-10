"""
Importador flexible de presupuestos Excel.

Soporta dos formatos:
  - Formato S10 clásico: columnas Ítem | Descripción | Und | Metrado | P.U.
  - Formato propio (Conila/oferta): Hoja1 con col A=código, B=descripción,
    C=unidad, D=metrado, E=P.U., F=subtotal

Detecta automáticamente la hoja más limpia y la fila de encabezados.
Soporta jerarquías de hasta 5 niveles.
"""

import io
import re
import openpyxl
import xlrd
from decimal import Decimal, InvalidOperation
from django.db import transaction
from .models import Partida, InsumoPresupuesto
from .pdf_parser import computar_parciales


# ── Utilidades ──────────────────────────────────────────────────

def _to_decimal(val):
    try:
        return Decimal(str(val)).quantize(Decimal('0.0001'))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _nivel_codigo(codigo):
    """
    Determina el nivel jerárquico a partir del código de ítem.
    Maneja:
      1, 2, 3          → nivel 1
      1.01, 1.02       → nivel 2  (float con 2 decimales)
      01.01.01         → nivel 3  (string con puntos)
      01.01.01.01      → nivel 4
      01.05.01.01.01   → nivel 5
    """
    if not codigo:
        return 1
    s = str(codigo).strip()
    # Intentar como número
    try:
        f = float(s)
        if f == int(f):
            return 1            # 1, 2, 3, 1.0, 2.0 → nivel 1
        return 2                # 1.1, 1.01, 1.10 → nivel 2
    except ValueError:
        pass
    # Formato string con puntos: 01.01.01, 01.01.01.01, 01.05.01.01.01
    partes = s.split('.')
    return len(partes)


def _es_fila_encabezado(row):
    """Devuelve True si la fila parece ser una cabecera de columnas."""
    textos = [str(c).strip().lower() for c in row if c is not None]
    keywords = {'item', 'partida', 'descripcion', 'descripción', 'und', 'unidad', 'metrado', 'cant.', 'pu', 'p.u.', 'precio'}
    return bool(keywords & set(textos))


def _detectar_columnas(row):
    """
    A partir de la fila de encabezado, devuelve un dict con los índices
    de columna para cada campo: codigo, nombre, unidad, cantidad, precio_unitario.
    Si no encuentra alguna columna, usa el offset posicional desde la primera col con datos.
    """
    mapping = {}
    for i, cell in enumerate(row):
        if cell is None:
            continue
        t = str(cell).strip().lower()
        if t in ('item', 'ítem', 'partida', 'cod', 'código', 'codigo') and 'codigo' not in mapping:
            mapping['codigo'] = i
        elif t in ('descripcion', 'descripción', 'descripcion.', 'nombre', 'descripcion de la partida') and 'nombre' not in mapping:
            mapping['nombre'] = i
        elif t in ('und', 'unid.', 'unid', 'unidad', 'u.m.') and 'unidad' not in mapping:
            mapping['unidad'] = i
        elif t in ('metrado', 'cant.', 'cant', 'cantidad', 'medida') and 'cantidad' not in mapping:
            mapping['cantidad'] = i
        elif t in ('pu', 'p.u.', 'precio', 'precio unit.', 'p. unit.', 'p.unitario',
                   'precio unitario', 'p. unitario', 'costo unit.', 'costo unitario', 'c.u.') and 'precio_unitario' not in mapping:
            mapping['precio_unitario'] = i

    # Fallback posicional desde la primera columna con datos
    base = None
    for i, cell in enumerate(row):
        if cell is not None:
            base = i
            break
    if base is not None:
        mapping.setdefault('codigo',          base)
        mapping.setdefault('nombre',          base + 1)
        mapping.setdefault('unidad',          base + 2)
        mapping.setdefault('cantidad',        base + 3)
        mapping.setdefault('precio_unitario', base + 4)
    return mapping


def _es_codigo_valido(val):
    """Determina si un valor puede ser un código de ítem (no un header ni texto largo)."""
    if val is None:
        return False
    s = str(val).strip()
    if not s or len(s) > 25:
        return False
    # Número entero
    try:
        f = float(s)
        return f > 0
    except ValueError:
        pass
    # Patrón de código: dígitos y puntos únicamente (01.01.01.01)
    return bool(re.match(r'^\d+(\.\d+)+$', s))


def _elegir_hoja(wb):
    """Elige la hoja más limpia para importar el presupuesto."""
    preferidas = ['hoja1', 'presupuesto', 'partidas', 'budget']
    for nombre in wb.sheetnames:
        if nombre.lower().strip() in preferidas:
            return wb[nombre]
    return wb.active


# ── Importador principal ─────────────────────────────────────────

def importar_presupuesto_excel(archivo, presupuesto):
    """
    Lee un archivo .xlsx/.xls e importa las partidas al presupuesto dado.
    Retorna (total_partidas, info_dict) donde info_dict tiene estadísticas del import.
    """
    nombre = getattr(archivo, 'name', '').lower()

    if nombre.endswith('.xls'):
        filas = _leer_xls(archivo)
    else:
        filas = _leer_xlsx(archivo)

    if not filas:
        raise ValueError('El archivo está vacío o no tiene filas de datos.')

    with transaction.atomic():
        presupuesto.partidas.all().delete()

        pila = {}
        orden = 0
        total = 0
        con_precio = 0

        for fila in filas:
            codigo = str(fila.get('codigo', '') or '').strip()
            nombre_p = str(fila.get('nombre', '') or '').strip()

            if not codigo or not nombre_p:
                continue
            if _es_fila_encabezado([codigo, nombre_p]):
                continue
            if not _es_codigo_valido(codigo):
                continue

            nivel = _nivel_codigo(codigo)
            padre = pila.get(nivel - 1)
            pu = _to_decimal(fila.get('precio_unitario'))

            partida = Partida.objects.create(
                presupuesto=presupuesto,
                padre=padre,
                codigo=codigo,
                nombre=nombre_p,
                nivel=nivel,
                orden=orden,
                unidad=str(fila.get('unidad', '') or '').strip()[:20],
                cantidad=_to_decimal(fila.get('cantidad')),
                precio_unitario=pu,
            )
            pila[nivel] = partida
            for k in list(pila.keys()):
                if k > nivel:
                    del pila[k]
            orden += 1
            total += 1
            if pu > 0:
                con_precio += 1

    computar_parciales(presupuesto)
    info = {
        'total':      total,
        'con_precio': con_precio,
        'sin_precio': total - con_precio,
    }
    return total, info


# ── Detección automática de tipo ────────────────────────────────

_PALABRAS_INSUMO = {
    'MANO DE OBRA', 'MATERIALES', 'MATERIAL', 'EQUIPO', 'EQUIPOS',
    'MAQUINARIA', 'SUB-CONTRATOS', 'SUBCONTRATOS', 'SUBCONTRATO',
    'MANO DE OBRA.',
}

def detectar_tipo_excel(archivo):
    """
    Lee las primeras 60 filas del Excel y determina si es:
      'presupuesto' — códigos jerárquicos S10 (01, 01.01, 01.01.01)
      'insumos'     — lista plana de recursos con códigos de catálogo
    También resetea el puntero del archivo para que pueda leerse de nuevo.
    """
    nombre = getattr(archivo, 'name', '').lower()
    contenido = archivo.read()
    try:
        archivo.seek(0)
    except Exception:
        pass

    if nombre.endswith('.xls'):
        wb  = xlrd.open_workbook(file_contents=contenido)
        ws  = wb.sheet_by_index(0)
        raw = [[ws.cell_value(r, c) for c in range(ws.ncols)]
               for r in range(min(ws.nrows, 60))]
    else:
        wb  = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws  = wb.active
        raw = [list(row) for _, row in zip(range(60), ws.iter_rows(values_only=True))]
        wb.close()

    score_insumo      = 0
    score_presupuesto = 0

    for row in raw:
        # Cabecera de categoría → señal fuerte de insumos
        for celda in row:
            if celda and str(celda).strip().upper() in _PALABRAS_INSUMO:
                score_insumo += 5

        # Analizar el primer campo con datos que parezca código
        for celda in row:
            if celda is None or str(celda).strip() == '':
                continue
            s = str(celda).strip()
            try:
                f = float(s)
                if f > 10000:           # código de catálogo grande (ej: 470050565)
                    score_insumo += 1
                elif 0 < f <= 999:      # código raíz jerárquico (01, 02, 03...)
                    score_presupuesto += 1
            except ValueError:
                pass
            # patrón jerárquico: 01.01, 01.01.01, 01.05.03.01
            if re.match(r'^\d{1,3}(\.\d{2,3}){1,4}$', s):
                score_presupuesto += 3
            break   # solo evaluar la primera celda con datos

    return 'insumos' if score_insumo > score_presupuesto else 'presupuesto'


def importar_automatico(archivo, presupuesto):
    """
    Detecta el tipo del Excel e importa al módulo correcto.
    Retorna (tipo_detectado, total, info_dict).
    """
    tipo = detectar_tipo_excel(archivo)

    if tipo == 'insumos':
        total = importar_insumos_excel(archivo, presupuesto)
        info  = {'total': total, 'con_precio': total, 'sin_precio': 0}
    else:
        total, info = importar_presupuesto_excel(archivo, presupuesto)

    return tipo, total, info


# ── Importador de insumos ────────────────────────────────────────

TIPO_MAP = {
    'MANO DE OBRA':  'MANO_OBRA',
    'MANO DE OBRA.': 'MANO_OBRA',
    'MATERIALES':    'MATERIAL',
    'MATERIAL':      'MATERIAL',
    'EQUIPO':        'EQUIPO',
    'EQUIPOS':       'EQUIPO',
    'MAQUINARIA':    'EQUIPO',
    'SUB-CONTRATOS': 'SUBCONTRATO',
    'SUBCONTRATOS':  'SUBCONTRATO',
    'SUBCONTRATO':   'SUBCONTRATO',
}

# Fallback de normalización cuando la BD no está disponible
_UNIDAD_NORM_FALLBACK = {
    'hh': 'HH', 'hm': 'HM', 'he': 'HE',
    'm2': 'm2', 'm²': 'm2', 'm3': 'm3', 'm³': 'm3',
    'ml': 'ml', 'mll': 'ml', 'kg': 'kg', 'tn': 'tn', 'ton': 'tn',
    'glb': 'glb', 'gb': 'glb', 'gl': 'glb',
    'und': 'und', 'un': 'und', 'u': 'und',
    'dia': 'dia', 'día': 'dia', 'mes': 'mes', 'sem': 'sem',
    'lt': 'lt', 'lts': 'lt', 'p2': 'p2', 'p3': 'p3',
}


def _cargar_mapa_unidades():
    """Construye {alias_lower: codigo} desde UnidadMedida en BD. Llama una vez por importación."""
    mapa = {}
    try:
        from apps.configuracion.models import UnidadMedida
        for u in UnidadMedida.objects.filter(activo=True).only('codigo', 'aliases'):
            mapa[u.codigo.lower()] = u.codigo
            if u.aliases:
                for alias in u.aliases.split(','):
                    a = alias.strip().lower()
                    if a:
                        mapa[a] = u.codigo
    except Exception:
        pass
    return mapa or _UNIDAD_NORM_FALLBACK


def _normalizar_unidad(valor, mapa=None):
    """Normaliza la unidad de medida usando el mapa BD o el fallback hardcodeado."""
    if not valor:
        return ''
    s = str(valor).strip()
    if mapa is None:
        mapa = _UNIDAD_NORM_FALLBACK
    return mapa.get(s.lower(), s)


# Aliases de cabecera para detección automática de columnas
_COL_ALIASES = {
    'codigo':      {'código', 'codigo', 'cód.', 'cod.', 'cod', 'ítem', 'item', 'nro', 'nro.', 'n°', 'id'},
    'descripcion': {'descripción', 'descripcion', 'recurso', 'nombre', 'descripcion del recurso',
                    'denominación', 'denominacion', 'detalle', 'material', 'insumo'},
    'unidad':      {'unidad', 'und', 'und.', 'unid', 'unid.', 'unidades',
                    'u.m.', 'u.m', 'um', 'u/m', 'unidad de medida'},
    'cantidad':    {'cantidad', 'cántidad', 'cant.', 'cant', 'cantid.', 'ctd', 'ctd.',
                    'metrado', 'cuadrilla', 'cuad.'},
    'precio':      {'precio', 'precio s/.', 'precio s/', 'precio unitario', 'precio compra',
                    'p.u.', 'p.u', 'p/u', 'pu',
                    'costo', 'costo unitario', 'costo unit.', 'c.u.', 'c.u',
                    'p. compra', 'p.compra', 'valor'},
    'total':       {'total', 'total s/.', 'total s/', 'parcial', 'importe'},
}


def _detectar_columnas_insumos(raw_rows):
    """
    Busca la fila de cabecera en las primeras 20 filas y devuelve un dict
    {campo: índice_columna}. Si no encuentra cabecera usa posiciones por defecto.
    """
    for row in raw_rows[:20]:
        celdas = [str(c).strip().lower() if c is not None else '' for c in row]
        mapa = {}
        for campo, aliases in _COL_ALIASES.items():
            for i, celda in enumerate(celdas):
                if celda in aliases:
                    mapa[campo] = i
                    break
        # Cabecera válida si encontramos al menos código + descripción
        if 'codigo' in mapa and 'descripcion' in mapa:
            return mapa
    # Fallback: columnas fijas C-G (índices 2-6)
    return {'codigo': 2, 'descripcion': 3, 'unidad': 4, 'cantidad': 5, 'precio': 6, 'total': 7}


def _cargar_mapa_cargos():
    """Lee CargoManoObra de DB y construye dict {variante_lower: codigo}."""
    try:
        from apps.configuracion.models import CargoManoObra
        mapa = {}
        for cargo in CargoManoObra.objects.filter(activo=True):
            if cargo.variantes:
                for v in cargo.variantes.split(','):
                    v = v.strip().lower()
                    if v:
                        mapa[v] = cargo.codigo
        return mapa
    except Exception:
        return {}


def _codigo_cargo(descripcion, mapa_cargos=None):
    """Devuelve el código del cargo si la descripción coincide con alguna variante, o None."""
    if mapa_cargos is None:
        mapa_cargos = _cargar_mapa_cargos()
    desc = descripcion.lower()
    for variante, cod in mapa_cargos.items():
        if variante in desc:
            return cod
    return None


def importar_insumos_excel(archivo, presupuesto):
    """
    Lee el archivo de insumos detectando las columnas automáticamente por cabecera.
    Soporta: Código, Descripción, Und, Precio, Total  (sin Cantidad)
         y:  Código, Descripción, Und, Cantidad, Precio, Total
    Las filas sin código válido pero con texto en TIPO_MAP marcan la categoría.
    Los correlativos se asignan en orden alfabético por descripción dentro de cada tipo.
    """
    nombre = getattr(archivo, 'name', '').lower()
    contenido = archivo.read()
    if nombre.endswith('.xls'):
        wb = xlrd.open_workbook(file_contents=contenido)
        ws = wb.sheet_by_index(0)
        raw_rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    else:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

    cols = _detectar_columnas_insumos(raw_rows)
    mapa_unidades = _cargar_mapa_unidades()
    mapa_cargos   = _cargar_mapa_cargos()
    n_cargos_db   = len({v for v in mapa_cargos.values()}) if mapa_cargos else 4

    presupuesto.insumos.all().delete()

    # ── Primera pasada: recolectar filas agrupadas por tipo ──────────
    tipo_actual = 'MATERIAL'
    grupos      = {}   # {tipo: [item_dict, ...]}
    tipo_orden  = []   # orden de aparición de tipos en el archivo
    desconocidas = set()

    for row in raw_rows:
        def get(idx):
            try: return row[idx]
            except (IndexError, TypeError): return None

        cod  = get(cols.get('codigo',  2))
        desc = get(cols.get('descripcion', 3))

        # Fila de categoría: sin código válido → buscar cabecera de tipo en toda la fila
        if not _es_codigo_valido(cod):
            for celda in row:
                if celda:
                    texto = str(celda).strip().upper()
                    if texto in TIPO_MAP:
                        tipo_actual = TIPO_MAP[texto]
                        if tipo_actual not in grupos:
                            grupos[tipo_actual] = []
                            tipo_orden.append(tipo_actual)
                        break
            continue

        if not desc:
            continue

        und = get(cols.get('unidad', 4))
        und_raw = str(und).strip() if und else ''
        if und_raw and und_raw.lower() not in mapa_unidades:
            desconocidas.add(und_raw)

        precio    = _to_decimal(get(cols.get('precio', cols.get('cantidad', 6))))
        total_val = _to_decimal(get(cols.get('total', 7)))

        if 'cantidad' in cols and 'precio' in cols:
            cant     = _to_decimal(get(cols['cantidad']))
            precio   = _to_decimal(get(cols['precio']))
            total_calc = (cant * precio).quantize(Decimal('0.01'))
            total_val  = _to_decimal(get(cols['total'])) if 'total' in cols else Decimal('0')
            total_val  = total_val or total_calc
        else:
            cant = Decimal('0')
            if total_val and precio:
                cant = (total_val / precio).quantize(Decimal('0.0001')) if precio else Decimal('0')

        if tipo_actual not in grupos:
            grupos[tipo_actual] = []
            tipo_orden.append(tipo_actual)

        grupos[tipo_actual].append({
            'desc':     str(desc).strip(),
            'und_raw':  und_raw,
            'und':      und,
            'cant':     cant,
            'precio':   precio,
            'total':    total_val,
        })

    # ── Segunda pasada: ordenar alfabéticamente y guardar ────────────
    total = 0
    with transaction.atomic():
        for tipo in tipo_orden:
            items = grupos[tipo]

            if tipo == 'MANO_OBRA':
                # Cargos conocidos mantienen su código fijo; desconocidos van al final ordenados
                conocidos   = [(it, _codigo_cargo(it['desc'], mapa_cargos)) for it in items
                               if _codigo_cargo(it['desc'], mapa_cargos) is not None]
                desconocidos = sorted(
                    [it for it in items if _codigo_cargo(it['desc'], mapa_cargos) is None],
                    key=lambda x: x['desc'].lower(),
                )
                # Ordenar conocidos por su código fijo
                conocidos.sort(key=lambda x: x[1])
                n = n_cargos_db
                ordered_final = [(it, cod) for it, cod in conocidos]
                for it in desconocidos:
                    n += 1
                    ordered_final.append((it, str(n)))
            else:
                items_sorted  = sorted(items, key=lambda x: x['desc'].lower())
                ordered_final = [(it, str(i)) for i, it in enumerate(items_sorted, 1)]

            for it, cod_asignado in ordered_final:
                InsumoPresupuesto.objects.create(
                    presupuesto=presupuesto,
                    tipo=tipo,
                    codigo=cod_asignado,
                    descripcion=it['desc'][:400],
                    unidad=_normalizar_unidad(it['und'], mapa_unidades)[:20],
                    cantidad=it['cant'],
                    costo_unitario=it['precio'],
                    total=it['total'],
                )
                total += 1

    return total, sorted(desconocidas)


# ── Leer archivos ────────────────────────────────────────────────

def _leer_xlsx(archivo):
    contenido = io.BytesIO(archivo.read())
    wb = openpyxl.load_workbook(contenido, read_only=True, data_only=True)
    ws = _elegir_hoja(wb)
    raw = list(ws.iter_rows(values_only=True))
    wb.close()

    filas = []
    cols = None
    data_iniciada = False

    for row in raw:
        if not any(c is not None for c in row):
            continue
        if _es_fila_encabezado(row) and not data_iniciada:
            cols = _detectar_columnas(row)
            data_iniciada = True
            continue
        if not data_iniciada:
            continue
        filas.append(_mapear_fila(row, cols))

    # Sin encabezado: usar todas las filas con offset 0
    if not filas:
        for row in raw:
            if any(c is not None for c in row):
                filas.append(_mapear_fila(row, None))
    return filas


def _leer_xls(archivo):
    contenido = archivo.read()
    wb = xlrd.open_workbook(file_contents=contenido)
    ws = wb.sheet_by_index(0)
    filas = []
    cols = None
    data_iniciada = False
    for i in range(ws.nrows):
        row = [ws.cell_value(i, c) for c in range(ws.ncols)]
        if not any(row):
            continue
        if _es_fila_encabezado(row) and not data_iniciada:
            cols = _detectar_columnas(row)
            data_iniciada = True
            continue
        if not data_iniciada:
            continue
        filas.append(_mapear_fila(row, cols))
    return filas


def _mapear_fila(row, cols):
    if not cols:
        cols = {'codigo': 0, 'nombre': 1, 'unidad': 2, 'cantidad': 3, 'precio_unitario': 4}

    def get(key):
        idx = cols.get(key)
        if idx is None:
            return ''
        try:
            v = row[idx]
            return v if v is not None else ''
        except IndexError:
            return ''

    return {
        'codigo':          get('codigo'),
        'nombre':          get('nombre'),
        'unidad':          get('unidad'),
        'cantidad':        get('cantidad'),
        'precio_unitario': get('precio_unitario'),
    }


# Mantener compatibilidad con el nombre anterior
importar_s10_excel = importar_presupuesto_excel
