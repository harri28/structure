import openpyxl
from .models import Producto

HOJA_CATEGORIA = {
    'MATERIALES EN BD': 'MATERIAL',
    'EQUIPOS EN BD': 'EQUIPO',
    'EPP EN BD': 'EPP',
    'UTILES EN BD': 'UTIL',
    'HERRAMIENTAS EN BD': 'HERRAMIENTA',
    'OTROS EN BD': 'OTRO',
}


def importar_catalogo_excel(archivo):
    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    creados = 0
    actualizados = 0

    for nombre_hoja, categoria in HOJA_CATEGORIA.items():
        if nombre_hoja not in wb.sheetnames:
            continue
        ws = wb[nombre_hoja]
        for row in ws.iter_rows(min_row=8, values_only=True):
            codigo = str(row[1]).strip() if row[1] else ''
            descripcion = str(row[2]).strip() if row[2] else ''
            if not codigo or not descripcion or codigo == 'None' or codigo == 'CODIGO':
                continue
            obj, created = Producto.objects.update_or_create(
                codigo=codigo,
                defaults={'descripcion': descripcion, 'categoria': categoria},
            )
            if created:
                creados += 1
            else:
                actualizados += 1

    wb.close()
    return creados, actualizados
